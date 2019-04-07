import imapclient
import datetime
import pyzmail
import openpyxl
# make a connection telling that you want to connect to gmail with an ssl encryption
connection = imapclient.IMAPClient("imap.gmail.com", ssl = True)
# Logging into your gmail account
usrName = input("Enter your Username or Email : ")
password = input("Enter your password : ")
connection.login(usrName, password)
connection.select_folder("INBOX", readonly = True)
uniIds = connection.search([u'SINCE', datetime.date(2015, 8, 1)]) #date format YY-MM-DD
# Getting the emails (START)
emails = []
for uniId in uniIds:
	raw_msg = connection.fetch([uniId], ["BODY[]", "FLAGS"])
	msg = pyzmail.PyzMessage.factory(raw_msg[uniId][b"BODY[]"])
	if msg.get_addresses("from") not in emails:
		emails.append(msg.get_addresses("from"))

connection.logout()
# Getting the emails (END)

# Making The excel file (START)
wb = openpyxl.Workbook()
sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
row = 1
for email in emails:
	name = email[0][0]
	emailID = email[0][1]
	c1 = sheet.cell(row = row, column = 1)
	c2 = sheet.cell(row = row, column = 2)
	c1.value = name
	c2.value = emailID
	row += 1
wb.save("EmailBack.xlsx")
# Making The excel file (END)
	

